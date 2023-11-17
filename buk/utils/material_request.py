import frappe
from frappe.utils.data import nowdate
from frappe.utils.xlsxutils import make_xlsx
import six
from io import BytesIO
import openpyxl
from openpyxl.styles import Font

@frappe.whitelist()
def create_mr_from_quotation(quotation):
    quotation_doc = frappe.get_doc("Quotation", quotation)
    items = []
    for item in quotation_doc.items:
        items.append({
            'item_code': item.item_code,
            'item_name': item.item_name,
            'description': item.description,
            'item_group': item.item_group,
            'qty': item.qty,
            'uom': item.uom,
            'schedule_date': nowdate(),
            'rate': item.rate,
            'quotation': quotation
        })
    mr = frappe.get_doc({
        'doctype': 'Material Request',
        'material_request_type': 'Purchase',
        'schedule_date': nowdate(),
        'items': items
    }).insert()
    return mr.name

@frappe.whitelist()
def create_mr_excel_export(material_request):
	#get data from Material Request
	data = frappe.db.sql("""
		SELECT `sinvitem`.`qty`,
			`sinvitem`.`uom`,
			`sinvitem`.`item_code`,
			`sinvitem`.`item_group`
		FROM `tabMaterial Request Item` AS `sinvitem`
		LEFT JOIN `tabMaterial Request` AS `sinv` ON `sinvitem`.`parent` = `sinv`.`name`
		WHERE `sinv`.`name` = '{mr}'
		""".format(mr=material_request), as_dict=True)
	#create excel document	
	wb = openpyxl.Workbook(write_only=False)
	#create sheet in document
	ws = wb.create_sheet('MaterialRequest', 0)
	#create header row
	header = {"Menge": 8, "Einheit": 8, "Artikel-Code": 20, "Artikelgruppe": 30}
	#set width of columns and header to bold
	for col_num, (header, width) in enumerate(header.items(), 1):
		cell = ws.cell(row=1, column=col_num, value=header)
		cell.font = Font(bold=True)
		ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    #create rows
	for entry in data:
		row = [entry.get('qty'), entry.get('uom'), entry.get('item_code'), entry.get('item_group')]
		ws.append(row)
	#transform file to BytesIO and save file
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	#create doc and attach it to document
	file_data = xlsx_file.getvalue()
	file_entry = frappe.get_doc({
		'doctype': 'File',
		'file_name': 'Material_Request_Export_{0}.xlsx'.format(material_request),
		'folder': 'Home/Attachments',
		'is_privat': 1,
		'content': file_data,
		'attached_to_doctype': 'Material Request',
		'attached_to_name': material_request
	}).save()
	
	return
